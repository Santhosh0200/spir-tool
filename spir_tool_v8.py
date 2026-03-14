#!/usr/bin/env python3
"""
SPIR Extraction Tool  —  v8.4  (Unified Multi-Format Engine)

Duplicate detection rules:
  TYPE 1 — "Duplicate N"         : same DESCRIPTION + same PART NUMBER, SAP is same or absent
  TYPE 2 — "SAP NUMBER MISMATCH" : same DESCRIPTION + same PART NUMBER, but SAP numbers differ

Supported formats:
  FORMAT 1 — Multi-Annexure SPIR          (.xlsx with Annexure sheets)
  FORMAT 2 — Single-Sheet, one tag        (.xlsm, single tag column)
  FORMAT 3 — Single-Sheet, multi-tag      (.xlsm, multiple tag columns, flag per row)
  FORMAT 4 — Matrix SPIR + Continuation   (.xlsx, qty matrix + single continuation sheet)
  FORMAT 5 — Flag SPIR + Multi-Continuation (.xlsm, flag-per-row main + multiple
              continuation sheets each with many tag columns, item# links back to main)

FORMAT 5 layout:
  Main sheet  : Few tag cols (C3+), item data in fixed columns (C7=ITEM, C8=QTY, C9=DESC …)
                Per-row tag assignment = flag value (1) in the tag col of that row
  Cont sheets : Many tag cols starting at C4, item# in C3, item_ref in last used col (REMARKS)
                Each data row lists which tag cols are flagged — links item to main sheet for details
                Tag metadata: R4=MFR MODEL, R6=MFR SERIAL, R7=EQPT QTY per tag col

NEW DESCRIPTION logic (all formats):
  = Description + ", " + Part Number + ", " + Supplier
  (only non-empty parts are joined; if both absent → just Description)

Run:  python spir_tool_v8.py   →   http://localhost:5050
"""

import io, re, threading, uuid, webbrowser
from collections import defaultdict, OrderedDict
from flask import Flask, request, send_file, jsonify
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)

# ─── OUTPUT COLUMNS ───────────────────────────────────────────────────────────
OUTPUT_COLS = [
    'SPIR NO',
    'TAG NO',
    'EQPT MAKE',
    'EQPT MODEL',
    'EQPT SR NO',
    'EQPT QTY',
    'QUANTITY IDENTICAL PARTS FITTED',
    'ITEM NUMBER',
    'DESCRIPTION OF PARTS',
    'NEW DESCRIPTION OF PARTS',
    'DWG NO INCL POSN NO',
    'MANUFACTURER PART NUMBER',
    'SUPPLIER OCM NAME',
    'CURRENCY',
    'UNIT PRICE',
    'DELIVERY TIME IN WEEKS',
    'MIN MAX STOCK LVLS QTY',
    'UNIT OF MEASURE',
    'SAP NUMBER',
    'CLASSIFICATION OF PARTS',
    'DUPLICATE ID',
    'SHEET',
    'SPIR TYPE',
]

CI = {name: i for i, name in enumerate(OUTPUT_COLS)}

# ─── HELPERS ──────────────────────────────────────────────────────────────────

def cv(ws, r, c):
    v = ws.cell(r, c).value
    return str(v).strip() if v is not None else ''

def cn(ws, r, c):
    """Read a numeric cell. Returns int if whole number, float if fractional, None if empty."""
    v = ws.cell(r, c).value
    if v is None:
        return None
    try:
        f = float(v)
        return int(f) if f == int(f) else f
    except (ValueError, TypeError):
        return None

def ne(s):
    """Return s if non-empty, else None."""
    return s if s else None

def clean_num(v):
    """Convert a raw numeric value to int if whole number, float if fractional, None if invalid."""
    if v is None:
        return None
    try:
        f = float(v)
        return int(f) if f == int(f) else f
    except (ValueError, TypeError):
        return None

def norm(s):
    """Normalise string for comparison."""
    return re.sub(r'\s+', ' ', (s or '').strip().lower())

def make_new_desc(desc, mfr_part_no, supplier_name):
    """
    Concatenate Description + Part Number + Supplier.
    Rules:
      - Skip any part that is empty or None
      - Skip any part that is a placeholder: TBA, NA, N/A, TBC, "-", ".", "N.A", etc.
        (case-insensitive, stripped)
      - If part number AND supplier are both absent/placeholder → return Description only
    """
    # Placeholder values that should be treated as "not available"
    _PLACEHOLDERS = {'tba', 'na', 'n/a', 'n.a', 'n.a.', 'tbc', '-', '.', 'nil',
                     'none', 'not applicable', 'not available', 'unknown', '—', 'n.a'}

    def is_real(val):
        if not val:
            return False
        s = str(val).strip()
        return bool(s) and s.lower() not in _PLACEHOLDERS

    parts = [desc] if is_real(desc) else []
    if is_real(mfr_part_no):
        parts.append(str(mfr_part_no).strip())
    if is_real(supplier_name):
        parts.append(str(supplier_name).strip())

    return ', '.join(parts)

# ─── SPIR TYPE DETECTION ──────────────────────────────────────────────────────

def detect_spir_type(ws, checkbox_col):
    type_map = {
        2: 'Commissioning Spares',
        3: 'Initial Spares',
        4: 'Normal Operating Spares',
        5: 'Life Cycle Spares',
    }
    for row_idx, label in type_map.items():
        raw = ws.cell(row_idx, checkbox_col).value
        if raw is True or raw == 1 or str(raw).strip().lower() in ('true', '1', 'x', 'yes'):
            return label
    sn = ws.title.lower()
    if 'commission' in sn: return 'Commissioning Spares'
    if 'initial'   in sn: return 'Initial Spares'
    if 'normal'    in sn: return 'Normal Operating Spares'
    if 'life'      in sn: return 'Life Cycle Spares'
    return None

# ─── DUPLICATE DETECTION ──────────────────────────────────────────────────────

def compute_duplicate_ids(items):
    sig_map = defaultdict(list)
    for idx, item in enumerate(items):
        sig = (norm(item.get('desc')), norm(item.get('mfr_part_no')))
        sig_map[sig].append((idx, (item.get('sap_no') or '').strip()))

    labels = [''] * len(items)
    dup_counter = 1

    for sig, entries in sig_map.items():
        if len(entries) < 2:
            continue
        sap_values = [sap for _, sap in entries]
        non_empty_saps = [s for s in sap_values if s]
        unique_saps = set(non_empty_saps)
        if len(unique_saps) > 1:
            for idx, _ in entries:
                labels[idx] = 'SAP NUMBER MISMATCH'
        else:
            label = f'Duplicate {dup_counter}'
            dup_counter += 1
            for idx, _ in entries:
                labels[idx] = label

    return labels

# ─── FORMAT DETECTOR ─────────────────────────────────────────────────────────

_HDR_KEYWORDS = ('spare', 'record', 'spir', 'qatar', 'energy', 'number',
                 'interchangeability', 'authority', 'note')

def _is_tag_value(s):
    s = s.strip()
    if not s:
        return False
    sl = s.lower()
    return not any(kw in sl for kw in _HDR_KEYWORDS)

def _find_tag_cols(ms):
    tag_cols = []
    for c in range(3, ms.max_column + 1):
        v = (ms.cell(1, c).value or '')
        s = str(v).strip()
        if _is_tag_value(s):
            tag_cols.append(c)
        else:
            if tag_cols:
                break
    return tag_cols

def detect_format(wb):
    """
    Detection order:
      FORMAT 5 — multiple continuation sheets  (2+ sheets with 'continuation' in name)
      FORMAT 4 — single continuation sheet     (exactly 1 sheet with 'continuation')
      FORMAT 1 — annexure sheets present
      FORMAT 2/3 — single-sheet, count tag cols to differentiate
    """
    sheet_names_lower = [n.lower() for n in wb.sheetnames]
    cont_sheets = [n for n in sheet_names_lower if 'continuation' in n]

    if len(cont_sheets) >= 2:
        return 'FORMAT5'
    if len(cont_sheets) == 1:
        return 'FORMAT4'
    if any('annexure' in n for n in sheet_names_lower):
        return 'FORMAT1'

    main_name = next((n for n in wb.sheetnames if 'main' in n.lower()), wb.sheetnames[0])
    ms = wb[main_name]
    tag_cols = _find_tag_cols(ms)
    return 'FORMAT2' if len(tag_cols) <= 1 else 'FORMAT3'


# ─── DYNAMIC COLUMN FINDER ────────────────────────────────────────────────────
# Used by FORMAT 4 to locate data columns from header row dynamically

def find_data_col(header_cells, *keywords):
    """
    Search header_cells (list of (col_idx, value_str)) for the first cell
    whose lowercase value contains ALL of the given keywords.
    Returns the col_idx (1-based) or None.
    """
    for col_idx, val in header_cells:
        vl = val.lower()
        if all(k in vl for k in keywords):
            return col_idx
    return None

def parse_header_cells(ws, header_row):
    """Return list of (col_idx, str_value) for non-empty cells in header_row."""
    result = []
    for ci in range(1, ws.max_column + 1):
        v = ws.cell(header_row, ci).value
        if v is not None and str(v).strip():
            result.append((ci, str(v).strip()))
    return result


# ─── FORMAT 4 — MATRIX SPIR + CONTINUATION SHEET ─────────────────────────────
#
# MAIN SHEET layout:
#   Row 1        : C2="EQUIPMENT TAG No", C3+ = tag numbers (until non-tag col)
#   Row 2        : C21="EQUIPMENT:", C24=equipment_name
#   Row 3        : C21="MANUFACTURER:", C25=manufacturer
#   Row 4        : C21="SUPPLIER:", C23=supplier
#   Row 1,C21    : "25 SPIR NUMBER:", C25=spir_no
#   Row 6 (hdr)  : Column headers for spare part data
#   Row 7        : "No. OF UNITS" + qty per tag column
#   Row 8+       : Spare items — qty per tag col (C3..Cn), then C7=ITEM#, C8=QTY_IDENTICAL,
#                  C9=DESC, C10=DWG, C11=MFR_PART, C12=SUP_PART, C15=SUPPLIER,
#                  C21=CURRENCY, C22=PRICE, C23=DELIVERY, C24=MIN_MAX, C25=UOM,
#                  C26=SAP, C27=CLASSIFICATION
#
# CONTINUATION SHEET layout:
#   Row 1        : C2="EQUIPMENT TAG No", C3+ = MORE tag numbers
#   Row 7        : "No. OF UNITS" per tag col + C17="DESCRIPTION OF PARTS", C18="REMARKS"
#   Row 8+       : qty per tag col (C3..Cn), C16=seq, C17=desc, C18=item_ref (→ main ITEM#)
#
# Logic:
#   1. Parse main sheet → build item_map {item_num → spare_part_details}
#   2. Parse main sheet → build tag_item_qty {tag_no → {item_num → qty}}
#   3. Parse continuation sheet → build more tag_item_qty using item_ref to look up details
#   4. For each tag × item (where qty > 0): emit one equipment header + one detail row

def extract_format4(wb):
    # ── Identify sheets ──────────────────────────────────────────────────────
    main_name = next(
        (n for n in wb.sheetnames if 'spir' in n.lower() and 'continuation' not in n.lower()),
        wb.sheetnames[0]
    )
    cont_name = next(
        (n for n in wb.sheetnames if 'continuation' in n.lower()),
        None
    )
    ms = wb[main_name]
    cs = wb[cont_name] if cont_name else None

    # ── Metadata from main sheet ─────────────────────────────────────────────
    # SPIR NO: scan rows 1-5 for pattern or label
    spir_no = ''
    for ri in range(1, 6):
        for ci in range(1, ms.max_column + 1):
            cell_val = cv(ms, ri, ci)
            if 'spir number' in cell_val.lower() or 'spir no' in cell_val.lower():
                # value is in next non-empty cell on same row
                for cj in range(ci + 1, ms.max_column + 1):
                    v = cv(ms, ri, cj)
                    if v:
                        spir_no = v.split('\n')[0].strip()
                        break
                if spir_no:
                    break
            # Also try direct VEN- pattern
            if re.match(r'^VEN-\d{4}-', cell_val):
                spir_no = cell_val.split('\n')[0].strip()
        if spir_no:
            break

    equipment = ''
    manufacturer = ''
    supplier = ''
    for ri in range(1, 8):
        for ci in range(1, ms.max_column + 1):
            cell_val = cv(ms, ri, ci).lower().rstrip(':')
            if cell_val == 'equipment':
                for cj in range(ci + 1, ms.max_column + 1):
                    v = cv(ms, ri, cj)
                    if v:
                        equipment = v; break
            elif cell_val == 'manufacturer':
                for cj in range(ci + 1, ms.max_column + 1):
                    v = cv(ms, ri, cj)
                    if v:
                        manufacturer = v; break
            elif cell_val == 'supplier':
                for cj in range(ci + 1, ms.max_column + 1):
                    v = cv(ms, ri, cj)
                    if v:
                        supplier = v; break

    # SPIR type — look for checkbox col or sheet name
    spir_type = detect_spir_type(ms, 30)
    if not spir_type:
        spir_type = detect_spir_type(ms, 28)

    # ── Find tag columns in main sheet (row 1, starting C3) ─────────────────
    def get_tag_cols_from_sheet(ws):
        tag_cols = []   # list of (col_idx, tag_no)
        for ci in range(3, ws.max_column + 1):
            v = cv(ws, 1, ci)
            if not v:
                break
            if _is_tag_value(v):
                tag_cols.append((ci, v))
            else:
                if tag_cols:
                    break
        return tag_cols

    main_tag_cols = get_tag_cols_from_sheet(ms)
    # eqpt_model from row 4, eqpt_sr from row 6 per tag col
    def get_tag_meta(ws, tag_cols):
        meta = {}
        for ci, tag_no in tag_cols:
            meta[tag_no] = {
                'col':        ci,
                'tag_no':     tag_no,
                'eqpt_model': cv(ws, 4, ci) or None,
                'eqpt_sr':    cv(ws, 6, ci) or None,
                'eqpt_qty':   cn(ws, 7, ci),
            }
        return meta

    main_tag_meta = get_tag_meta(ms, main_tag_cols)
    main_tag_col_idx = [ci for ci, _ in main_tag_cols]

    # ── Dynamically find data column indices from header row ─────────────────
    # Header row = first row (>=row 5) where ITEM NUMBER or DESCRIPTION appears
    hdr_row = None
    for ri in range(4, min(12, ms.max_row + 1)):
        row_str = ' '.join(str(ms.cell(ri, ci).value or '').lower() for ci in range(1, ms.max_column + 1))
        if 'item number' in row_str or 'description of parts' in row_str:
            hdr_row = ri
            break
    if hdr_row is None:
        hdr_row = 6  # fallback

    hdr_cells = parse_header_cells(ms, hdr_row)

    def dc(*keywords):
        return find_data_col(hdr_cells, *keywords)

    col_item     = dc('item number')
    col_qty_id   = dc('total no. of identical')
    col_desc     = dc('description of parts')
    col_dwg      = dc('dwg no')
    col_mfr_part = dc('manufacturer part number')
    col_sup_part = dc('suppliers part number')
    col_supplier = dc('supplier/ocm name')
    col_currency = dc('currency')
    col_price    = dc('unit price')
    col_delivery = dc('delivery time')
    col_minmax   = dc('min/max stock')
    col_uom      = dc('unit  of measure') or dc('unit of measure')
    col_sap      = dc('sap number')
    col_class    = dc('classification of parts')

    # ── Parse main sheet spare items → item_map ──────────────────────────────
    # Data starts after header row + 1 sub-header row
    data_start = hdr_row + 2
    item_map = {}     # item_num (int) → dict of spare part fields
    raw_items_main = []

    # Also track per-tag quantities from main: {tag_no: {item_num: qty}}
    main_tag_qtys = {tag_no: {} for _, tag_no in main_tag_cols}

    for ri in range(data_start, ms.max_row + 1):
        item_val = cn(ms, ri, col_item) if col_item else None
        desc_val = cv(ms, ri, col_desc) if col_desc else ''
        if item_val is None or not desc_val or len(desc_val) < 2:
            continue
        # Stop at footer rows
        if desc_val.lower().startswith(('project', 'company', 'engineering', 'reminder', 'technical')):
            break

        item_num = int(item_val)
        mfr_part = ne(cv(ms, ri, col_mfr_part)) if col_mfr_part else None
        sup_part = ne(cv(ms, ri, col_sup_part)) if col_sup_part else None
        supp_name = ne(cv(ms, ri, col_supplier)) if col_supplier else None
        sap_val = ne(cv(ms, ri, col_sap)) if col_sap else None

        spare = {
            'item_num':       item_num,
            'desc':           desc_val,
            'dwg_no':         ne(cv(ms, ri, col_dwg)) if col_dwg else None,
            'mfr_part_no':    mfr_part or sup_part,
            'supplier_name':  supp_name,
            'currency':       ne(cv(ms, ri, col_currency)) if col_currency else None,
            'unit_price':     cn(ms, ri, col_price) if col_price else None,
            'delivery':       ne(cv(ms, ri, col_delivery)) if col_delivery else None,
            'min_max':        ne(cv(ms, ri, col_minmax)) if col_minmax else None,
            'uom':            ne(cv(ms, ri, col_uom)) if col_uom else None,
            'sap_no':         sap_val,
            'classification': ne(cv(ms, ri, col_class)) if col_class else None,
            'qty_identical':  cn(ms, ri, col_qty_id) if col_qty_id else None,
        }
        item_map[item_num] = spare
        raw_items_main.append(spare)

        # Record per-tag qty from matrix
        for ci, tag_no in main_tag_cols:
            q = cn(ms, ri, ci)
            if q is not None and q > 0:
                main_tag_qtys[tag_no][item_num] = q

    # ── Parse continuation sheet ─────────────────────────────────────────────
    cont_tag_cols = []
    cont_tag_meta = {}
    cont_tag_qtys = {}

    if cs is not None:
        cont_tag_cols = get_tag_cols_from_sheet(cs)
        cont_tag_meta = get_tag_meta(cs, cont_tag_cols)
        cont_tag_qtys = {tag_no: {} for _, tag_no in cont_tag_cols}

        # Find header row in continuation sheet
        cont_hdr_row = None
        for ri in range(4, min(12, cs.max_row + 1)):
            row_str = ' '.join(str(cs.cell(ri, ci).value or '').lower()
                               for ci in range(1, cs.max_column + 1))
            if 'description of parts' in row_str or 'no. of units' in row_str:
                cont_hdr_row = ri
                break
        if cont_hdr_row is None:
            cont_hdr_row = 7

        cont_data_start = cont_hdr_row + 1

        # Detect continuation sheet column positions dynamically
        # item_ref col = column with "REMARKS" header (or last numeric col before desc)
        # desc col = column with "DESCRIPTION OF PARTS" header
        # seq col = column just before desc col
        cont_hdr_cells = parse_header_cells(cs, cont_hdr_row)
        cont_col_desc = find_data_col(cont_hdr_cells, 'description of parts')
        cont_col_ref  = find_data_col(cont_hdr_cells, 'remarks')
        # seq col = cont_col_ref - 1 usually, but let's just read cont_col_desc - 1
        cont_col_seq  = (cont_col_desc - 1) if cont_col_desc and cont_col_desc > 1 else None

        # Identify last tag column (first col with non-tag content after tag block)
        last_tag_ci = max(ci for ci, _ in cont_tag_cols) if cont_tag_cols else 15

        for ri in range(cont_data_start, cs.max_row + 1):
            # item ref is in cont_col_ref
            item_ref_val = cn(cs, ri, cont_col_ref) if cont_col_ref else None
            if item_ref_val is None:
                # Try reading as int from cont_col_ref
                raw = cs.cell(ri, cont_col_ref).value if cont_col_ref else None
                try:
                    item_ref_val = int(raw) if raw is not None else None
                except (ValueError, TypeError):
                    item_ref_val = None

            if item_ref_val is None:
                continue

            item_num = int(item_ref_val)

            # Record per-tag qty
            for ci, tag_no in cont_tag_cols:
                q = cn(cs, ri, ci)
                if q is not None and q > 0:
                    cont_tag_qtys[tag_no][item_num] = q

    # ── Duplicate detection across all items ─────────────────────────────────
    dup_ids = compute_duplicate_ids(raw_items_main)
    dup_map = {item['item_num']: dup_ids[i] for i, item in enumerate(raw_items_main)}

    # ── Build output rows ────────────────────────────────────────────────────
    out_rows = []
    sheet_label = main_name.upper().strip()

    def make_hdr_row(tag_no, eqpt_model, eqpt_sr, eqpt_qty, source_sheet):
        row = [None] * len(OUTPUT_COLS)
        row[CI['SPIR NO']]              = spir_no
        row[CI['TAG NO']]               = tag_no
        row[CI['EQPT MAKE']]            = manufacturer
        row[CI['EQPT MODEL']]           = eqpt_model
        row[CI['EQPT SR NO']]           = eqpt_sr
        row[CI['EQPT QTY']]             = int(eqpt_qty) if eqpt_qty else None
        row[CI['DESCRIPTION OF PARTS']] = equipment
        row[CI['DUPLICATE ID']]         = 0
        row[CI['SHEET']]                = source_sheet
        row[CI['SPIR TYPE']]            = spir_type
        return row

    def make_det_row(tag_no, eqpt_model, eqpt_sr, spare, qty_for_tag, dup_id, source_sheet):
        new_desc = make_new_desc(spare['desc'], spare['mfr_part_no'], spare['supplier_name'])
        row = [None] * len(OUTPUT_COLS)
        row[CI['SPIR NO']]                         = spir_no
        row[CI['TAG NO']]                          = tag_no
        row[CI['EQPT MAKE']]                       = manufacturer
        row[CI['EQPT MODEL']]                      = eqpt_model
        row[CI['EQPT SR NO']]                      = eqpt_sr
        row[CI['QUANTITY IDENTICAL PARTS FITTED']] = qty_for_tag
        row[CI['ITEM NUMBER']]                     = spare['item_num']
        row[CI['DESCRIPTION OF PARTS']]            = spare['desc']
        row[CI['NEW DESCRIPTION OF PARTS']]        = new_desc
        row[CI['DWG NO INCL POSN NO']]             = spare['dwg_no']
        row[CI['MANUFACTURER PART NUMBER']]        = spare['mfr_part_no']
        row[CI['SUPPLIER OCM NAME']]               = spare['supplier_name']
        row[CI['CURRENCY']]                        = spare['currency']
        row[CI['UNIT PRICE']]                      = spare['unit_price']
        row[CI['DELIVERY TIME IN WEEKS']]          = spare['delivery']
        row[CI['MIN MAX STOCK LVLS QTY']]          = spare['min_max']
        row[CI['UNIT OF MEASURE']]                 = spare['uom']
        row[CI['SAP NUMBER']]                      = spare['sap_no']
        row[CI['CLASSIFICATION OF PARTS']]         = spare['classification']
        row[CI['DUPLICATE ID']]                    = dup_id if dup_id else 0
        row[CI['SHEET']]                           = source_sheet
        row[CI['SPIR TYPE']]                       = spir_type
        return row

    # Emit main sheet tags
    for ci, tag_no in main_tag_cols:
        tmeta  = main_tag_meta[tag_no]
        eq_qty = tmeta['eqpt_qty']
        out_rows.append(make_hdr_row(tag_no, tmeta['eqpt_model'], tmeta['eqpt_sr'],
                                     eq_qty, sheet_label))
        for item_num, qty in sorted(main_tag_qtys[tag_no].items()):
            spare  = item_map.get(item_num)
            if not spare:
                continue
            dup_id = dup_map.get(item_num, '')
            out_rows.append(make_det_row(tag_no, tmeta['eqpt_model'], tmeta['eqpt_sr'],
                                         spare, qty, dup_id, sheet_label))

    # Emit continuation sheet tags
    cont_label = (cont_name or 'CONTINUATION SHEET').upper().strip()
    for ci, tag_no in cont_tag_cols:
        tmeta  = cont_tag_meta[tag_no]
        eq_qty = tmeta['eqpt_qty']
        out_rows.append(make_hdr_row(tag_no, tmeta['eqpt_model'], tmeta['eqpt_sr'],
                                     eq_qty, cont_label))
        for item_num, qty in sorted(cont_tag_qtys[tag_no].items()):
            spare  = item_map.get(item_num)
            if not spare:
                continue
            dup_id = dup_map.get(item_num, '')
            out_rows.append(make_det_row(tag_no, tmeta['eqpt_model'], tmeta['eqpt_sr'],
                                         spare, qty, dup_id, cont_label))

    total_tags = len(main_tag_cols) + len(cont_tag_cols)

    return {
        'format':         'FORMAT4',
        'spir_no':        spir_no,
        'equipment':      equipment,
        'manufacturer':   manufacturer,
        'supplier':       supplier,
        'spir_type':      spir_type,
        'eqpt_qty':       total_tags,
        'spare_items':    len(item_map),
        'total_tags':     total_tags,
        'annexure_count': 0,
        'annexure_stats': {
            f'{main_name} (Main)': len(main_tag_cols),
            f'{cont_name} (Continuation)': len(cont_tag_cols),
        } if cont_name else {f'{main_name} (Main)': len(main_tag_cols)},
        'rows':           out_rows,
    }


# ─── FORMAT 1 — MULTI-ANNEXURE ────────────────────────────────────────────────

def extract_format1(wb):
    main_name = next((n for n in wb.sheetnames if 'main' in n.lower()), None)
    if not main_name:
        raise ValueError('No MAIN SHEET found.')
    ms = wb[main_name]

    spir_no      = cv(ms, 1, 27).rstrip()
    equipment    = cv(ms, 2, 26)
    manufacturer = cv(ms, 3, 27)
    supplier     = cv(ms, 4, 25)
    spir_type    = detect_spir_type(ms, 30)

    if not spir_no:
        for ri in range(1, 6):
            for ci in range(1, ms.max_column + 1):
                s = cv(ms, ri, ci)
                if re.match(r'^VEN-\d{4}-', s):
                    spir_no = s; break
            if spir_no: break

    ann_unit_counts = {}
    for ann_idx, col in enumerate(range(3, 9), 1):
        n = cn(ms, 7, col)
        if n and n > 0:
            ann_unit_counts[ann_idx] = int(n)
    total_eqpt_qty = sum(ann_unit_counts.values()) if ann_unit_counts else None

    raw_items = []
    for ri in range(8, ms.max_row + 1):
        item_num_val = cn(ms, ri, 9)
        desc         = cv(ms, ri, 11)
        if item_num_val is None or not desc or len(desc) < 3:
            continue
        ann_flags = [cn(ms, ri, col) not in (None, 0) for col in range(3, 9)]
        mfr_part  = ne(cv(ms, ri, 13))
        supp_name = ne(cv(ms, ri, 17))
        raw_items.append({
            'item_num':       int(item_num_val),
            'qty_identical':  1,
            'desc':           desc,
            'dwg_no':         ne(cv(ms, ri, 12)),
            'mfr_part_no':    mfr_part,
            'supplier_name':  supp_name,
            'currency':       ne(cv(ms, ri, 23)),
            'unit_price':     cn(ms, ri, 24),
            'delivery':       cn(ms, ri, 25),
            'min_max':        ne(cv(ms, ri, 26)),
            'uom':            ne(cv(ms, ri, 27)),
            'sap_no':         ne(cv(ms, ri, 28)),
            'classification': ne(cv(ms, ri, 29)),
            'ann_flags':      ann_flags,
        })

    dup_ids = compute_duplicate_ids(raw_items)

    ann_sheets     = [n for n in wb.sheetnames if 'annexure' in n.lower()]
    annexure_data  = {}
    annexure_stats = {}
    for idx, sname in enumerate(ann_sheets):
        ann_num = idx + 1
        ws = wb[sname]
        hdr_row = None
        for ri in range(1, min(ws.max_row + 1, 10)):
            s = '|'.join(str(ws.cell(ri, ci).value or '') for ci in range(1, 10)).lower()
            if 'sr. no' in s or 'sr.no' in s:
                hdr_row = ri; break
        tags = []
        if hdr_row:
            for ri in range(hdr_row + 2, ws.max_row + 1):
                sr = cv(ws, ri, 2)
                if not sr or not sr.replace('.', '').isdigit():
                    continue
                tags.append({
                    'valve_tag':  cv(ws, ri, 5),
                    'mfr_model':  cv(ws, ri, 24),
                    'mfr_serial': cv(ws, ri, 23),
                })
        annexure_data[ann_num]             = tags
        annexure_stats[f'Annexure {ann_num}'] = len(tags)

    out_rows    = []
    sheet_label = main_name.upper()

    for item_idx, item in enumerate(raw_items):
        dup_id = dup_ids[item_idx]
        applicable_anns = [i + 1 for i, f in enumerate(item['ann_flags']) if f]
        if not applicable_anns:
            applicable_anns = list(annexure_data.keys()) or [1]

        new_desc = make_new_desc(item['desc'], item['mfr_part_no'], item['supplier_name'])

        for ann_num in applicable_anns:
            tags         = annexure_data.get(ann_num, [])
            eqpt_qty_ann = ann_unit_counts.get(ann_num, total_eqpt_qty)

            def hdr(tag_no=None, eqpt_mdl=None, eqpt_sr=None):
                row = [None] * len(OUTPUT_COLS)
                row[CI['SPIR NO']]              = spir_no
                row[CI['TAG NO']]               = tag_no
                row[CI['EQPT MAKE']]            = manufacturer
                row[CI['EQPT MODEL']]           = eqpt_mdl
                row[CI['EQPT SR NO']]           = eqpt_sr
                row[CI['EQPT QTY']]             = eqpt_qty_ann
                row[CI['DESCRIPTION OF PARTS']] = equipment
                row[CI['DUPLICATE ID']]         = 0
                row[CI['SHEET']]                = sheet_label
                row[CI['SPIR TYPE']]            = spir_type
                return row

            def det(tag_no=None, eqpt_mdl=None, eqpt_sr=None):
                row = [None] * len(OUTPUT_COLS)
                row[CI['SPIR NO']]                         = spir_no
                row[CI['TAG NO']]                          = tag_no
                row[CI['EQPT MAKE']]                       = manufacturer
                row[CI['EQPT MODEL']]                      = eqpt_mdl
                row[CI['EQPT SR NO']]                      = eqpt_sr
                row[CI['QUANTITY IDENTICAL PARTS FITTED']] = item['qty_identical']
                row[CI['ITEM NUMBER']]                     = item['item_num']
                row[CI['DESCRIPTION OF PARTS']]            = item['desc']
                row[CI['NEW DESCRIPTION OF PARTS']]        = new_desc
                row[CI['DWG NO INCL POSN NO']]             = item['dwg_no']
                row[CI['MANUFACTURER PART NUMBER']]        = item['mfr_part_no']
                row[CI['SUPPLIER OCM NAME']]               = item['supplier_name']
                row[CI['CURRENCY']]                        = item['currency']
                row[CI['UNIT PRICE']]                      = item['unit_price']
                row[CI['DELIVERY TIME IN WEEKS']]          = item['delivery']
                row[CI['MIN MAX STOCK LVLS QTY']]          = item['min_max']
                row[CI['UNIT OF MEASURE']]                 = item['uom']
                row[CI['SAP NUMBER']]                      = item['sap_no']
                row[CI['CLASSIFICATION OF PARTS']]         = item['classification']
                row[CI['DUPLICATE ID']]                    = dup_id if dup_id else 0
                row[CI['SHEET']]                           = sheet_label
                row[CI['SPIR TYPE']]                       = spir_type
                return row

            if not tags:
                out_rows.append(hdr())
                continue

            out_rows.append(hdr())
            out_rows.append(det())
            for tag in tags:
                tn = ne(tag['valve_tag'])
                em = ne(tag['mfr_model'])
                es = ne(tag['mfr_serial'])
                out_rows.append(hdr(tn, em, es))
                out_rows.append(det(tn, em, es))

    return {
        'format':         'FORMAT1',
        'spir_no':        spir_no,
        'equipment':      equipment,
        'manufacturer':   manufacturer,
        'supplier':       supplier,
        'spir_type':      spir_type,
        'eqpt_qty':       total_eqpt_qty,
        'spare_items':    len(raw_items),
        'total_tags':     sum(len(v) for v in annexure_data.values()),
        'annexure_count': len(ann_sheets),
        'annexure_stats': annexure_stats,
        'rows':           out_rows,
    }


# ─── FORMAT 2 — SINGLE-SHEET, SINGLE TAG (.xlsm) ────────────────────────────

def extract_format2(wb):
    main_name = next(
        (n for n in wb.sheetnames if 'main' in n.lower()),
        wb.sheetnames[0]
    )
    ms = wb[main_name]

    tag_cols = []
    for c in range(3, ms.max_column + 1):
        v = cv(ms, 1, c)
        if v and not any(kw in v.lower() for kw in ('spare', 'record', 'spir', 'qatar', 'number')):
            tag_cols.append(c)
        elif tag_cols:
            break
    if not tag_cols:
        tag_cols = [3]

    tags_meta = []
    for c in tag_cols:
        tags_meta.append({
            'tag_no':     cv(ms, 1, c) or None,
            'eqpt_model': cv(ms, 4, c) or None,
            'eqpt_sr':    cv(ms, 6, c) or None,
        })

    spir_no      = cv(ms, 1, 25).rstrip()
    equipment    = cv(ms, 2, 24)
    manufacturer = cv(ms, 3, 25)
    supplier     = cv(ms, 4, 23)
    spir_type    = detect_spir_type(ms, 28)

    eqpt_qty = 0
    for c in tag_cols:
        n = cn(ms, 7, c)
        if n and n > 0:
            eqpt_qty += int(n)
    if not eqpt_qty:
        eqpt_qty = len(tag_cols)

    if not spir_no:
        for ri in range(1, 6):
            for ci_s in range(1, ms.max_column + 1):
                s = cv(ms, ri, ci_s)
                if re.match(r'^VEN-\d{4}-', s):
                    spir_no = s; break
            if spir_no: break

    raw_items = []
    for ri in range(8, ms.max_row + 1):
        item_num_val = cn(ms, ri, 7)
        desc         = cv(ms, ri, 9)
        if item_num_val is None or not desc or len(desc) < 3:
            continue
        if desc.lower().startswith(('project', 'company', 'engineering', 'reminder')):
            break
        mfr_part = ne(cv(ms, ri, 11))
        sup_part = ne(cv(ms, ri, 12))
        supp_name = ne(cv(ms, ri, 15))
        qty_val  = cn(ms, ri, 8) or cn(ms, ri, 3)
        raw_items.append({
            'item_num':       int(item_num_val),
            'qty_identical':  qty_val,
            'desc':           desc,
            'dwg_no':         ne(cv(ms, ri, 10)),
            'mfr_part_no':    mfr_part or sup_part,
            'supplier_name':  supp_name,
            'currency':       ne(cv(ms, ri, 21)),
            'unit_price':     cn(ms, ri, 22),
            'delivery':       cn(ms, ri, 23),
            'min_max':        ne(cv(ms, ri, 24)),
            'uom':            ne(cv(ms, ri, 25)),
            'sap_no':         ne(cv(ms, ri, 26)),
            'classification': ne(cv(ms, ri, 27)),
        })

    dup_ids = compute_duplicate_ids(raw_items)

    out_rows    = []
    sheet_label = main_name.upper().strip()

    for t_idx, tmeta in enumerate(tags_meta):
        row = [None] * len(OUTPUT_COLS)
        row[CI['SPIR NO']]              = spir_no
        row[CI['TAG NO']]               = tmeta['tag_no']
        row[CI['EQPT MAKE']]            = manufacturer
        row[CI['EQPT MODEL']]           = tmeta['eqpt_model']
        row[CI['EQPT SR NO']]           = tmeta['eqpt_sr']
        row[CI['EQPT QTY']]             = eqpt_qty if t_idx == 0 else None
        row[CI['DESCRIPTION OF PARTS']] = equipment
        row[CI['DUPLICATE ID']]         = 0
        row[CI['SHEET']]                = sheet_label
        row[CI['SPIR TYPE']]            = spir_type
        out_rows.append(row)

    for item_idx, item in enumerate(raw_items):
        dup_id   = dup_ids[item_idx]
        new_desc = make_new_desc(item['desc'], item['mfr_part_no'], item['supplier_name'])

        for tmeta in tags_meta:
            row = [None] * len(OUTPUT_COLS)
            row[CI['SPIR NO']]                         = spir_no
            row[CI['TAG NO']]                          = tmeta['tag_no']
            row[CI['EQPT MAKE']]                       = manufacturer
            row[CI['EQPT MODEL']]                      = tmeta['eqpt_model']
            row[CI['EQPT SR NO']]                      = tmeta['eqpt_sr']
            row[CI['QUANTITY IDENTICAL PARTS FITTED']] = item['qty_identical']
            row[CI['ITEM NUMBER']]                     = item['item_num']
            row[CI['DESCRIPTION OF PARTS']]            = item['desc']
            row[CI['NEW DESCRIPTION OF PARTS']]        = new_desc
            row[CI['DWG NO INCL POSN NO']]             = item['dwg_no']
            row[CI['MANUFACTURER PART NUMBER']]        = item['mfr_part_no']
            row[CI['SUPPLIER OCM NAME']]               = item['supplier_name']
            row[CI['CURRENCY']]                        = item['currency']
            row[CI['UNIT PRICE']]                      = item['unit_price']
            row[CI['DELIVERY TIME IN WEEKS']]          = item['delivery']
            row[CI['MIN MAX STOCK LVLS QTY']]          = item['min_max']
            row[CI['UNIT OF MEASURE']]                 = item['uom']
            row[CI['SAP NUMBER']]                      = item['sap_no']
            row[CI['CLASSIFICATION OF PARTS']]         = item['classification']
            row[CI['DUPLICATE ID']]                    = dup_id if dup_id else 0
            row[CI['SHEET']]                           = sheet_label
            row[CI['SPIR TYPE']]                       = spir_type
            out_rows.append(row)

    return {
        'format':         'FORMAT2',
        'spir_no':        spir_no,
        'equipment':      equipment,
        'manufacturer':   manufacturer,
        'supplier':       supplier,
        'spir_type':      spir_type,
        'eqpt_qty':       eqpt_qty,
        'spare_items':    len(raw_items),
        'total_tags':     len(tags_meta),
        'annexure_count': 0,
        'annexure_stats': {},
        'rows':           out_rows,
    }


# ─── FORMAT 3 — SINGLE-SHEET, MULTIPLE TAGS (.xlsm) ─────────────────────────

def extract_format3(wb):
    main_name = next((n for n in wb.sheetnames if 'main' in n.lower()), wb.sheetnames[0])
    ms = wb[main_name]

    tag_cols  = _find_tag_cols(ms)
    tags_meta = {}
    for c in tag_cols:
        tag_no  = cv(ms, 1, c)
        model   = cv(ms, 4, c) or None
        serial  = cv(ms, 6, c) or None
        qty_raw = cn(ms, 7, c)
        qty     = int(qty_raw) if qty_raw and qty_raw > 0 else 1
        tags_meta[c] = {
            'col':    c,
            'tag_no': tag_no or None,
            'model':  model,
            'serial': serial,
            'qty':    qty,
        }

    spir_no      = cv(ms, 1, 25).split('\n')[0].strip()
    equipment    = cv(ms, 2, 24)
    manufacturer = cv(ms, 3, 25)
    supplier     = cv(ms, 4, 23)
    spir_type    = detect_spir_type(ms, 28)
    sheet_label  = main_name.upper().strip()

    if not spir_no:
        for ri in range(1, 6):
            for ci_s in range(1, ms.max_column + 1):
                s = cv(ms, ri, ci_s)
                if re.match(r'^VEN-\d{4}-', s):
                    spir_no = s; break
            if spir_no: break

    raw_items = []
    for ri in range(8, ms.max_row + 1):
        item_num_val = cn(ms, ri, 7)
        desc         = cv(ms, ri, 9)
        if item_num_val is None or not desc or len(desc) < 3:
            continue
        if desc.lower().startswith(('project', 'company', 'engineering', 'reminder')):
            break

        # Collect ALL tag columns that are flagged for this item row
        # (a single item can belong to multiple tags — value in that col = qty for that tag)
        flagged_cols = {}   # col_idx → qty_for_that_tag
        for c in tag_cols:
            v = ms.cell(ri, c).value
            try:
                fv = float(v) if v is not None else 0
                if fv > 0:
                    flagged_cols[c] = clean_num(v)
            except (ValueError, TypeError):
                pass

        if not flagged_cols:
            continue

        mfr_part  = ne(cv(ms, ri, 11))
        sup_part  = ne(cv(ms, ri, 12))
        supp_name = ne(cv(ms, ri, 15))
        qty_val   = cn(ms, ri, 8)

        base_item = {
            'item_num':       int(item_num_val),
            'qty_total':      qty_val,       # total across all tags (C8)
            'desc':           desc,
            'dwg_no':         ne(cv(ms, ri, 10)),
            'mfr_part_no':    mfr_part or sup_part,
            'supplier_name':  supp_name,
            'currency':       ne(cv(ms, ri, 21)),
            'unit_price':     cn(ms, ri, 22),
            'delivery':       cn(ms, ri, 23),
            'min_max':        ne(cv(ms, ri, 24)),
            'uom':            ne(cv(ms, ri, 25)),
            'sap_no':         ne(cv(ms, ri, 26)),
            'classification': ne(cv(ms, ri, 27)),
        }

        # Emit one entry per flagged tag column, carrying that tag's specific qty
        for col_c, tag_qty in flagged_cols.items():
            item_copy = dict(base_item)
            item_copy['tag_col']       = col_c
            item_copy['qty_identical'] = tag_qty   # qty from THIS tag's column
            raw_items.append(item_copy)

    # Duplicate detection: run on unique items (by item_num) to avoid falsely
    # flagging shared spares that legitimately appear under multiple tags.
    # Rule: SAP NUMBER MISMATCH only — same desc+part but different SAP numbers.
    # Same item shared across tags (same SAP) is NOT a duplicate.
    unique_items_by_num = {}
    for item in raw_items:
        n = item['item_num']
        if n not in unique_items_by_num:
            unique_items_by_num[n] = item
    unique_list = list(unique_items_by_num.values())
    unique_dup_ids = compute_duplicate_ids(unique_list)
    # Build item_num → dup_label, keeping only SAP MISMATCH labels
    item_dup_map = {}
    for i, item in enumerate(unique_list):
        label = unique_dup_ids[i]
        # Only propagate SAP NUMBER MISMATCH; drop generic "Duplicate N" across tags
        item_dup_map[item['item_num']] = label if label == 'SAP NUMBER MISMATCH' else ''

    out_rows = []
    tag_items = OrderedDict((c, []) for c in tag_cols)
    for item in raw_items:
        tag_items[item['tag_col']].append(item)

    for c, items_for_tag in tag_items.items():
        tmeta = tags_meta[c]

        hdr = [None] * len(OUTPUT_COLS)
        hdr[CI['SPIR NO']]              = spir_no
        hdr[CI['TAG NO']]               = tmeta['tag_no']
        hdr[CI['EQPT MAKE']]            = manufacturer
        hdr[CI['EQPT MODEL']]           = tmeta['model']
        hdr[CI['EQPT SR NO']]           = tmeta['serial']
        hdr[CI['EQPT QTY']]             = tmeta['qty']
        hdr[CI['DESCRIPTION OF PARTS']] = equipment
        hdr[CI['DUPLICATE ID']]         = 0
        hdr[CI['SHEET']]                = sheet_label
        hdr[CI['SPIR TYPE']]            = spir_type
        out_rows.append(hdr)

        for item in items_for_tag:
            dup_id   = item_dup_map.get(item['item_num'], '')
            new_desc = make_new_desc(item['desc'], item['mfr_part_no'], item['supplier_name'])

            det = [None] * len(OUTPUT_COLS)
            det[CI['SPIR NO']]                         = spir_no
            det[CI['TAG NO']]                          = tmeta['tag_no']
            det[CI['EQPT MAKE']]                       = manufacturer
            det[CI['EQPT MODEL']]                      = tmeta['model']
            det[CI['EQPT SR NO']]                      = tmeta['serial']
            det[CI['QUANTITY IDENTICAL PARTS FITTED']] = item['qty_identical']
            det[CI['ITEM NUMBER']]                     = item['item_num']
            det[CI['DESCRIPTION OF PARTS']]            = item['desc']
            det[CI['NEW DESCRIPTION OF PARTS']]        = new_desc
            det[CI['DWG NO INCL POSN NO']]             = item['dwg_no']
            det[CI['MANUFACTURER PART NUMBER']]        = item['mfr_part_no']
            det[CI['SUPPLIER OCM NAME']]               = item['supplier_name']
            det[CI['CURRENCY']]                        = item['currency']
            det[CI['UNIT PRICE']]                      = item['unit_price']
            det[CI['DELIVERY TIME IN WEEKS']]          = item['delivery']
            det[CI['MIN MAX STOCK LVLS QTY']]          = item['min_max']
            det[CI['UNIT OF MEASURE']]                 = item['uom']
            det[CI['SAP NUMBER']]                      = item['sap_no']
            det[CI['CLASSIFICATION OF PARTS']]         = item['classification']
            det[CI['DUPLICATE ID']]                    = dup_id if dup_id else 0
            det[CI['SHEET']]                           = sheet_label
            det[CI['SPIR TYPE']]                       = spir_type
            out_rows.append(det)

    total_tags = len(tag_cols)
    return {
        'format':         'FORMAT3',
        'spir_no':        spir_no,
        'equipment':      equipment,
        'manufacturer':   manufacturer,
        'supplier':       supplier,
        'spir_type':      spir_type,
        'eqpt_qty':       sum(t['qty'] for t in tags_meta.values()),
        'spare_items':    len(raw_items),
        'total_tags':     total_tags,
        'annexure_count': 0,
        'annexure_stats': {tags_meta[c]['tag_no']: len(tag_items[c]) for c in tag_cols},
        'rows':           out_rows,
    }


# ─── FORMAT 5 — FLAG SPIR + MULTIPLE CONTINUATION SHEETS (.xlsm) ─────────────
#
# MAIN SHEET layout (identical column positions to FORMAT 3/4):
#   Row 1       : C2="EQUIPMENT TAG No", C3+ = tag numbers
#   Row 2       : metadata (EQUIPMENT in fixed col)
#   Row 6       : column headers (ITEM NUMBER, DESCRIPTION OF PARTS, …)
#   Row 7       : "No. OF UNITS" + sub-headers
#   Row 8+      : spare items — C7=ITEM#, C8=QTY_IDENTICAL, C9=DESC, C10=DWG,
#                 C11=MFR_PART, C12=SUP_PART, C13=MAT_SPEC, C14=MAT_CERT,
#                 C15=SUPPLIER, C21=CURRENCY, C22=PRICE, C23=DELIVERY,
#                 C24=MIN_MAX, C25=UOM, C26=SAP, C27=CLASSIFICATION
#                 Tag flag = non-zero value in that tag's column for this item row
#
# CONTINUATION SHEET layout (each sheet):
#   Row 1       : C2="EQUIP'T / OR TAG", C4+ = tag numbers (stop at label col)
#   Row 4       : MFR MODEL per tag col
#   Row 6       : MFR SERIAL per tag col
#   Row 7       : "No. OF UNITS" per tag col; last label col = "REMARKS"
#   Row 8+      : C3=item_seq, flagged tag cols = non-zero, last col=item_ref (→ main ITEM#)
#
# KEY INSIGHT:
#   - Main sheet: each data row belongs to SPECIFIC tag(s) via flag value in tag col
#   - Cont sheet: C3=seq counter, last col (REMARKS) = item ref → main ITEM# for details
#   - Tag block in cont sheet starts at C4, ends before label column
#   - Dynamic: detect tag cols, item col, remarks col from each sheet independently

def _parse_cont_sheet_tags(ws):
    """
    Dynamically detect tag columns in a continuation sheet.
    Tags are in row 1 starting at C4, continuing until an empty or label cell.
    Returns list of (col_idx, tag_no) pairs.
    """
    tag_cols = []
    for ci in range(4, ws.max_column + 1):
        v = str(ws.cell(1, ci).value or '').strip()
        if not v:
            break  # stop at first empty
        # Stop if it looks like a label (e.g., "SPIR NUMBER:", "REMARKS")
        vl = v.lower()
        if any(kw in vl for kw in ('spir number', 'remarks', 'equipment:', 'manufacturer:', 'supplier:')):
            break
        if _is_tag_value(v):
            tag_cols.append((ci, v))
        # Don't stop on non-tag — could be a label mid-way; but empty does stop
    return tag_cols

def _find_remarks_col(ws, tag_cols):
    """
    Find the REMARKS column in a continuation sheet.
    It's the first column after the tag block (in row 7) that contains 'REMARKS'.
    Falls back to last_tag_col + 1.
    """
    last_tag_ci = max(ci for ci, _ in tag_cols) if tag_cols else 40
    for ci in range(last_tag_ci + 1, ws.max_column + 1):
        v = str(ws.cell(7, ci).value or '').strip().lower()
        if 'remark' in v:
            return ci
        v1 = str(ws.cell(1, ci).value or '').strip().lower()
        if 'remark' in v1:
            return ci
    return last_tag_ci + 1

def _parse_main_sheet_format5(ms):
    """
    Parse main sheet for FORMAT 5.
    Returns:
      metadata dict, item_map {item_num→spare_dict}, tag_cols [(ci, tag_no)],
      main_tag_meta {tag_no→{model,serial,qty}}, per_tag_items {tag_no→[item_num]}
    """
    # ── Metadata ────────────────────────────────────────────────────────────
    spir_no = ''
    equipment = ''
    manufacturer = ''
    supplier = ''
    spir_type = None

    for ri in range(1, 8):
        for ci in range(1, ms.max_column + 1):
            cell = str(ms.cell(ri, ci).value or '').strip()
            cl = cell.lower().rstrip(':')
            nv = ''  # next non-empty value on same row
            for cj in range(ci + 1, ms.max_column + 1):
                nv = str(ms.cell(ri, cj).value or '').strip()
                if nv:
                    break
            if 'spir number' in cl or re.match(r'^VEN-\d{4}-', cell):
                if re.match(r'^VEN-\d{4}-', cell):
                    spir_no = cell.split('\n')[0].strip()
                elif nv:
                    spir_no = nv.split('\n')[0].strip()
            elif cl == 'equipment':
                equipment = nv
            elif cl == 'manufacturer':
                manufacturer = nv
            elif cl == 'supplier':
                supplier = nv

    spir_type = detect_spir_type(ms, 28)
    if not spir_type:
        spir_type = detect_spir_type(ms, 30)

    # ── Tag columns in row 1 starting at C3 ─────────────────────────────────
    tag_cols = []
    for ci in range(3, ms.max_column + 1):
        v = str(ms.cell(1, ci).value or '').strip()
        if not v or not _is_tag_value(v):
            if tag_cols:
                break
            continue
        tag_cols.append((ci, v))

    main_tag_meta = {}
    for ci, tag_no in tag_cols:
        main_tag_meta[tag_no] = {
            'tag_no':     tag_no,
            'eqpt_model': str(ms.cell(4, ci).value or '').strip() or None,
            'eqpt_sr':    str(ms.cell(6, ci).value or '').strip() or None,
            'eqpt_qty':   cn(ms, 7, ci),
        }

    # ── Dynamically find data column positions from header row ───────────────
    hdr_row = None
    for ri in range(4, min(12, ms.max_row + 1)):
        row_str = ' '.join(str(ms.cell(ri, ci).value or '').lower()
                           for ci in range(1, ms.max_column + 1))
        if 'item number' in row_str or 'description of parts' in row_str:
            hdr_row = ri
            break
    if hdr_row is None:
        hdr_row = 6

    hdr_cells = parse_header_cells(ms, hdr_row)

    def dc(*kws):
        return find_data_col(hdr_cells, *kws)

    col_item     = dc('item number')     or 7
    col_qty_id   = dc('total no. of identical') or 8
    col_desc     = dc('description of parts')   or 9
    col_dwg      = dc('dwg no')          or 10
    col_mfr_part = dc('manufacturer part number') or 11
    col_sup_part = dc('suppliers part number')    or 12
    col_mat_spec = dc('material spec')   or 13
    col_mat_cert = dc('material cert')   or 14
    col_supplier = dc('supplier/ocm name')       or 15
    col_currency = dc('currency')        or 21
    col_price    = dc('unit price')      or 22
    col_delivery = dc('delivery time')   or 23
    col_minmax   = dc('min/max stock')   or 24
    col_uom      = dc('unit  of measure') or dc('unit of measure') or 25
    col_sap      = dc('sap number')      or 26
    col_class    = dc('classification of parts') or 27

    # ── Parse spare items + tag assignments ──────────────────────────────────
    data_start = hdr_row + 2
    item_map = {}
    raw_items = []
    # Key fix: track per-column, not per-tag-name (same tag name can appear in 2+ cols)
    per_col_items = {ci: [] for ci, _ in tag_cols}         # col_idx → [item_num]
    per_col_qty   = {}                                       # (col_idx, item_num) → flag qty

    for ri in range(data_start, ms.max_row + 1):
        item_val = cn(ms, ri, col_item)
        desc_val = str(ms.cell(ri, col_desc).value or '').strip()
        if item_val is None or not desc_val or len(desc_val) < 2:
            continue
        if desc_val.lower().startswith(('project', 'company', 'engineering', 'reminder', 'technical')):
            break

        item_num  = int(item_val)
        mfr_part  = ne(str(ms.cell(ri, col_mfr_part).value or '').strip())
        sup_part  = ne(str(ms.cell(ri, col_sup_part).value or '').strip())
        supp_name = ne(str(ms.cell(ri, col_supplier).value or '').strip())
        sap_val   = ne(str(ms.cell(ri, col_sap).value or '').strip())

        spare = {
            'item_num':       item_num,
            'desc':           desc_val,
            'dwg_no':         ne(str(ms.cell(ri, col_dwg).value or '').strip()),
            'mfr_part_no':    mfr_part or sup_part,
            'supplier_name':  supp_name,
            'currency':       ne(str(ms.cell(ri, col_currency).value or '').strip()),
            'unit_price':     cn(ms, ri, col_price),
            'delivery':       ne(str(ms.cell(ri, col_delivery).value or '').strip()),
            'min_max':        ne(str(ms.cell(ri, col_minmax).value or '').strip()),
            'uom':            ne(str(ms.cell(ri, col_uom).value or '').strip()),
            'sap_no':         sap_val,
            'classification': ne(str(ms.cell(ri, col_class).value or '').strip()),
        }
        item_map[item_num] = spare
        raw_items.append(spare)

        # Flag scanning: use the VALUE in the tag column as the qty_identical for that tag
        for ci, tag_no in tag_cols:
            v = ms.cell(ri, ci).value
            try:
                fv = float(v) if v is not None else 0.0
                if fv > 0:
                    per_col_items[ci].append(item_num)
                    per_col_qty[(ci, item_num)] = clean_num(v)
            except (ValueError, TypeError):
                pass

    return {
        'spir_no': spir_no, 'equipment': equipment,
        'manufacturer': manufacturer, 'supplier': supplier,
        'spir_type': spir_type,
        'tag_cols': tag_cols, 'main_tag_meta': main_tag_meta,
        'item_map': item_map, 'raw_items': raw_items,
        'per_col_items': per_col_items,
        'per_col_qty':   per_col_qty,
    }

def _parse_cont_sheet_format5(ws, item_map):
    """
    Parse one continuation sheet for FORMAT 5.
    Returns list of (col_idx, tag_no, tag_meta_dict, [item_nums], {item_num: qty})
    Tracks by column index so duplicate tag names are kept as separate entries.
    """
    tag_cols = _parse_cont_sheet_tags(ws)
    if not tag_cols:
        return []

    # Tag metadata per column (not per tag name — duplicate names are valid)
    tag_meta_by_col = {}
    for ci, tag_no in tag_cols:
        tag_meta_by_col[ci] = {
            'tag_no':     tag_no,
            'eqpt_model': str(ws.cell(4, ci).value or '').strip() or None,
            'eqpt_sr':    str(ws.cell(6, ci).value or '').strip() or None,
            'eqpt_qty':   cn(ws, 7, ci),
        }

    # Find the REMARKS column (item_ref → main sheet ITEM NUMBER)
    remarks_col = _find_remarks_col(ws, tag_cols)
    item_col    = 3  # item sequence number always in C3

    # Build per-column item list and qty
    per_col_items = {ci: [] for ci, _ in tag_cols}
    per_col_qty   = {}  # (ci, item_num) → qty from flag value

    data_start = 8
    for ri in range(data_start, ws.max_row + 1):
        seq_val = ws.cell(ri, item_col).value
        if seq_val is None:
            continue
        # item_ref in REMARKS col → maps to main sheet ITEM NUMBER
        ref_raw = ws.cell(ri, remarks_col).value
        try:
            item_ref = int(float(str(ref_raw).strip())) if ref_raw is not None else None
        except (ValueError, TypeError):
            item_ref = None

        if item_ref is None:
            try:
                item_ref = int(float(str(seq_val).strip()))
            except (ValueError, TypeError):
                continue

        if item_ref not in item_map:
            continue

        # Scan tag columns — flag value = qty_identical for that tag
        for ci, tag_no in tag_cols:
            v = ws.cell(ri, ci).value
            try:
                fv = float(v) if v is not None else 0.0
                if fv > 0:
                    per_col_items[ci].append(item_ref)
                    per_col_qty[(ci, item_ref)] = clean_num(v)
            except (ValueError, TypeError):
                pass

    return [
        (ci, tag_no, tag_meta_by_col[ci], per_col_items[ci], per_col_qty)
        for ci, tag_no in tag_cols
    ]


def extract_format5(wb):
    """
    FORMAT 5: Main sheet with flag-based tag assignment +
              Multiple continuation sheets each with many tag columns.
    """
    # ── Identify sheets ──────────────────────────────────────────────────────
    main_name = next(
        (n for n in wb.sheetnames if 'main' in n.lower()),
        wb.sheetnames[0]
    )
    cont_names = [n for n in wb.sheetnames if 'continuation' in n.lower()]
    ms = wb[main_name]
    sheet_label = main_name.upper().strip()

    # ── Parse main sheet ─────────────────────────────────────────────────────
    parsed = _parse_main_sheet_format5(ms)
    spir_no      = parsed['spir_no']
    equipment    = parsed['equipment']
    manufacturer = parsed['manufacturer']
    supplier     = parsed['supplier']
    spir_type    = parsed['spir_type']
    item_map     = parsed['item_map']
    raw_items    = parsed['raw_items']
    main_tag_cols    = parsed['tag_cols']        # [(col_idx, tag_no), ...]
    main_tag_meta    = parsed['main_tag_meta']   # {tag_no: meta}
    main_per_col     = parsed['per_col_items']   # {col_idx: [item_num]}
    main_per_col_qty = parsed['per_col_qty']     # {(col_idx, item_num): qty}

    # ── Duplicate detection ──────────────────────────────────────────────────
    dup_ids = compute_duplicate_ids(raw_items)
    dup_map = {item['item_num']: dup_ids[i] for i, item in enumerate(raw_items)}

    # ── Helper: emit rows for one tag (identified by col_idx for uniqueness) ─
    def emit_tag(tag_no, tmeta, item_nums, col_qty_map, col_idx, src_label):
        rows = []
        # Equipment header row
        hdr = [None] * len(OUTPUT_COLS)
        hdr[CI['SPIR NO']]              = spir_no
        hdr[CI['TAG NO']]               = tag_no
        hdr[CI['EQPT MAKE']]            = manufacturer
        hdr[CI['EQPT MODEL']]           = tmeta.get('eqpt_model')
        hdr[CI['EQPT SR NO']]           = tmeta.get('eqpt_sr')
        hdr[CI['EQPT QTY']]             = int(tmeta['eqpt_qty']) if tmeta.get('eqpt_qty') else 1
        hdr[CI['DESCRIPTION OF PARTS']] = equipment
        hdr[CI['DUPLICATE ID']]         = 0
        hdr[CI['SHEET']]                = src_label
        hdr[CI['SPIR TYPE']]            = spir_type
        rows.append(hdr)

        for item_num in item_nums:
            spare = item_map.get(item_num)
            if not spare:
                continue
            dup_id   = dup_map.get(item_num, '')
            new_desc = make_new_desc(spare['desc'], spare['mfr_part_no'], spare['supplier_name'])
            # qty_identical = value from the tag's column (flag qty), not total C8
            qty_for_tag = col_qty_map.get((col_idx, item_num), 1)

            det = [None] * len(OUTPUT_COLS)
            det[CI['SPIR NO']]                         = spir_no
            det[CI['TAG NO']]                          = tag_no
            det[CI['EQPT MAKE']]                       = manufacturer
            det[CI['EQPT MODEL']]                      = tmeta.get('eqpt_model')
            det[CI['EQPT SR NO']]                      = tmeta.get('eqpt_sr')
            det[CI['QUANTITY IDENTICAL PARTS FITTED']] = qty_for_tag
            det[CI['ITEM NUMBER']]                     = spare['item_num']
            det[CI['DESCRIPTION OF PARTS']]            = spare['desc']
            det[CI['NEW DESCRIPTION OF PARTS']]        = new_desc
            det[CI['DWG NO INCL POSN NO']]             = spare['dwg_no']
            det[CI['MANUFACTURER PART NUMBER']]        = spare['mfr_part_no']
            det[CI['SUPPLIER OCM NAME']]               = spare['supplier_name']
            det[CI['CURRENCY']]                        = spare['currency']
            det[CI['UNIT PRICE']]                      = spare['unit_price']
            det[CI['DELIVERY TIME IN WEEKS']]          = spare['delivery']
            det[CI['MIN MAX STOCK LVLS QTY']]          = spare['min_max']
            det[CI['UNIT OF MEASURE']]                 = spare['uom']
            det[CI['SAP NUMBER']]                      = spare['sap_no']
            det[CI['CLASSIFICATION OF PARTS']]         = spare['classification']
            det[CI['DUPLICATE ID']]                    = dup_id if dup_id else 0
            det[CI['SHEET']]                           = src_label
            det[CI['SPIR TYPE']]                       = spir_type
            rows.append(det)
        return rows

    # ── Build output rows ─────────────────────────────────────────────────────
    out_rows = []
    annexure_stats = {}

    # Main sheet tags (iterate by col to preserve order and handle duplicate names)
    for ci, tag_no in main_tag_cols:
        tmeta     = main_tag_meta[tag_no]
        item_nums = main_per_col.get(ci, [])
        out_rows.extend(emit_tag(tag_no, tmeta, item_nums, main_per_col_qty, ci, sheet_label))

    annexure_stats[f'{main_name} (Main)'] = len(main_tag_cols)

    # Continuation sheet tags
    for cname in cont_names:
        cs         = wb[cname]
        cont_label = cname.upper().strip()
        cont_tags  = _parse_cont_sheet_format5(cs, item_map)
        # cont_tags = [(ci, tag_no, tmeta, [item_nums], col_qty_map), ...]
        for ci, tag_no, tmeta, item_nums, col_qty_map in cont_tags:
            out_rows.extend(emit_tag(tag_no, tmeta, item_nums, col_qty_map, ci, cont_label))
        annexure_stats[cname] = len(cont_tags)

    total_tags = (len(main_tag_cols) +
                  sum(v for k, v in annexure_stats.items() if '(Main)' not in k))

    return {
        'format':         'FORMAT5',
        'spir_no':        spir_no,
        'equipment':      equipment,
        'manufacturer':   manufacturer,
        'supplier':       supplier,
        'spir_type':      spir_type,
        'eqpt_qty':       total_tags,
        'spare_items':    len(item_map),
        'total_tags':     total_tags,
        'annexure_count': len(cont_names),
        'annexure_stats': annexure_stats,
        'rows':           out_rows,
    }




def extract_spir(wb):
    fmt = detect_format(wb)
    if fmt == 'FORMAT5':
        return extract_format5(wb)
    elif fmt == 'FORMAT4':
        return extract_format4(wb)
    elif fmt == 'FORMAT1':
        return extract_format1(wb)
    elif fmt == 'FORMAT3':
        return extract_format3(wb)
    else:
        return extract_format2(wb)


# ─── BUILD OUTPUT XLSX ────────────────────────────────────────────────────────

def build_xlsx(out_rows, spir_no):
    wb2 = openpyxl.Workbook()
    ws  = wb2.active
    ws.title = 'SPIR Extraction'

    hdr_fill   = PatternFill('solid', fgColor='375623')
    hdr_font   = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    hdr_align  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin       = Side(style='thin', color='BFBFBF')
    hdr_border = Border(top=thin, bottom=thin, left=thin, right=thin)

    data_font   = Font(name='Calibri', size=10, color='000000')
    data_align  = Alignment(vertical='center', wrap_text=False)
    data_border = Border(top=thin, bottom=thin, left=thin, right=thin)
    white_fill  = PatternFill('solid', fgColor='FFFFFF')

    ws.append(OUTPUT_COLS)
    for cell in ws[1]:
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = hdr_border
    ws.row_dimensions[1].height = 30

    for row in out_rows:
        ws.append(row)
        ri = ws.max_row
        for cell in ws[ri]:
            cell.font      = data_font
            cell.fill      = white_fill
            cell.border    = data_border
            cell.alignment = data_align
        ws.row_dimensions[ri].height = 15

    col_widths = {
        'SPIR NO': 24, 'TAG NO': 22, 'EQPT MAKE': 28, 'EQPT MODEL': 24,
        'EQPT SR NO': 12, 'EQPT QTY': 10,
        'QUANTITY IDENTICAL PARTS FITTED': 12, 'ITEM NUMBER': 10,
        'DESCRIPTION OF PARTS': 50, 'NEW DESCRIPTION OF PARTS': 60,
        'DWG NO INCL POSN NO': 42, 'MANUFACTURER PART NUMBER': 36,
        'SUPPLIER OCM NAME': 28, 'CURRENCY': 24, 'UNIT PRICE': 12,
        'DELIVERY TIME IN WEEKS': 14, 'MIN MAX STOCK LVLS QTY': 14,
        'UNIT OF MEASURE': 14, 'SAP NUMBER': 16,
        'CLASSIFICATION OF PARTS': 20, 'DUPLICATE ID': 22,
        'SHEET': 22, 'SPIR TYPE': 26,
    }
    for ci_1, col_name in enumerate(OUTPUT_COLS, 1):
        ws.column_dimensions[get_column_letter(ci_1)].width = col_widths.get(col_name, 14)

    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb2.save(buf)
    buf.seek(0)
    return buf.read()


# ─── FLASK ────────────────────────────────────────────────────────────────────

_results = {}

@app.route('/')
def index():
    return HTML

@app.route('/extract', methods=['POST'])
def extract():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    f = request.files['file']
    try:
        wb     = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
        result = extract_spir(wb)
        xlsx   = build_xlsx(result['rows'], result['spir_no'])
        fid    = str(uuid.uuid4())
        fname  = (result['spir_no'] or 'SPIR').replace(' ', '') + '_SPIR_Extraction.xlsx'
        _results[fid] = (xlsx, fname)

        dup_col    = CI['DUPLICATE ID']
        dup1_count = sum(1 for r in result['rows']
                         if isinstance(r[dup_col], str) and r[dup_col].startswith('Duplicate'))
        sap_count  = sum(1 for r in result['rows']
                         if isinstance(r[dup_col], str) and r[dup_col] == 'SAP NUMBER MISMATCH')

        return jsonify({
            'format':         result['format'],
            'spir_no':        result['spir_no'],
            'equipment':      result['equipment'],
            'manufacturer':   result['manufacturer'],
            'supplier':       result['supplier'],
            'spir_type':      result['spir_type'],
            'eqpt_qty':       result['eqpt_qty'],
            'spare_items':    result['spare_items'],
            'total_tags':     result['total_tags'],
            'annexure_count': result['annexure_count'],
            'annexure_stats': result['annexure_stats'],
            'dup1_count':     dup1_count,
            'sap_count':      sap_count,
            'total_rows':     len(result['rows']),
            'preview_cols':   OUTPUT_COLS,
            'preview_rows':   result['rows'][:12],
            'file_id':        fid,
            'filename':       fname,
        })
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500

@app.route('/download/<fid>')
def download(fid):
    if fid not in _results:
        return 'Not found', 404
    data, fname = _results[fid]
    return send_file(
        io.BytesIO(data), as_attachment=True, download_name=fname,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


# ─── HTML UI ─────────────────────────────────────────────────────────────────

HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>SPIR Extraction Tool v7</title>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=JetBrains+Mono:wght@400;500&display=swap" rel="stylesheet">
<style>
:root{
  --bg:#F0F4F8;--surface:#FFFFFF;--border:#D1D9E0;--border2:#E8EDF2;
  --navy:#1F4E79;--navy2:#2E6CA6;--accent:#2E6CA6;
  --green:#1E6B3C;--green-bg:#E8F5EE;
  --yellow:#7F6000;--yellow-bg:#FFF9E6;--yellow-border:#F0D060;
  --red:#C00000;--red-bg:#FDF0EE;--red-border:#F0B0A0;
  --text:#1A2332;--muted:#5A6A7A;--muted2:#8A9BAB
}
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:'Inter',sans-serif;background:var(--bg);color:var(--text);min-height:100vh}
.wrap{max-width:960px;margin:0 auto;padding:0 24px}

header{background:var(--navy);padding:0;border-bottom:3px solid var(--navy2)}
.hdr-inner{display:flex;align-items:center;gap:16px;padding:20px 24px;max-width:960px;margin:0 auto}
.logo{width:44px;height:44px;background:rgba(255,255,255,.15);border-radius:10px;
  display:grid;place-items:center;font-family:'JetBrains Mono',monospace;
  font-weight:700;font-size:18px;color:#fff;flex-shrink:0}
.hdr-text h1{font-size:19px;font-weight:700;color:#fff;letter-spacing:-.01em}
.hdr-text p{font-size:11px;color:rgba(255,255,255,.65);margin-top:3px;letter-spacing:.03em}
.ver{margin-left:auto;background:rgba(255,255,255,.12);border:1px solid rgba(255,255,255,.2);
  color:rgba(255,255,255,.85);font-size:10px;padding:3px 10px;border-radius:4px;
  font-family:'JetBrains Mono',monospace;white-space:nowrap}

main{padding:32px 0 80px}

.stepper{display:flex;align-items:center;background:var(--surface);
  border:1px solid var(--border);border-radius:10px;padding:16px 24px;
  margin-bottom:24px;gap:4px}
.step{display:flex;align-items:center;gap:10px;flex:1}
.s-line{flex:1;height:1px;background:var(--border2);transition:background .3s;margin:0 6px}
.s-line.done{background:var(--navy2)}
.s-num{width:28px;height:28px;border-radius:50%;border:1.5px solid var(--border);
  display:grid;place-items:center;font-size:11px;font-weight:600;
  color:var(--muted);background:var(--bg);flex-shrink:0;transition:all .3s}
.step.active .s-num{border-color:var(--navy);background:var(--navy);color:#fff}
.step.done .s-num{border-color:var(--navy2);background:var(--navy2);color:#fff}
.s-lbl{font-size:11px;font-weight:500;color:var(--muted);letter-spacing:.02em}
.step.active .s-lbl,.step.done .s-lbl{color:var(--text)}

.card{background:var(--surface);border:1px solid var(--border);border-radius:10px;
  padding:28px;margin-bottom:20px}
.card-title{font-size:12px;font-weight:600;letter-spacing:.06em;text-transform:uppercase;
  color:var(--muted);margin-bottom:20px;display:flex;align-items:center;gap:8px}
.card-title::before{content:'';width:3px;height:14px;background:var(--navy);
  border-radius:2px;flex-shrink:0}

.drop-zone{border:2px dashed var(--border);border-radius:8px;
  padding:48px 32px;text-align:center;cursor:pointer;transition:all .25s}
.drop-zone:hover,.drop-zone.over{border-color:var(--navy2);background:#F5F8FC}
.drop-icon{font-size:40px;margin-bottom:14px;display:block;opacity:.7}
.drop-zone h3{font-size:15px;font-weight:600;margin-bottom:6px}
.drop-zone p{font-size:12px;color:var(--muted);line-height:1.6}
.drop-zone input{display:none}
.fmt-pills{display:flex;gap:8px;justify-content:center;margin-top:12px;flex-wrap:wrap}
.fpill{background:#EEF4FB;border:1px solid #C3D8EF;color:var(--navy2);
  font-size:10px;font-weight:500;padding:3px 10px;border-radius:20px}

.file-sel{display:none;align-items:center;gap:14px;
  background:#F0F7FF;border:1px solid #C3D8EF;border-radius:8px;
  padding:14px 18px;margin-top:14px}
.file-sel.show{display:flex;animation:slideIn .25s ease}
@keyframes slideIn{from{opacity:0;transform:translateY(-4px)}to{opacity:1;transform:translateY(0)}}
.fi-info{flex:1}
.fi-info strong{display:block;font-size:12.5px;color:var(--navy);font-weight:600}
.fi-info span{font-size:10.5px;color:var(--muted)}
.check-icon{color:#1E6B3C;font-size:16px;font-weight:700}

.btn-go{width:100%;padding:13px;margin-top:18px;background:var(--navy);
  border:none;border-radius:8px;color:#fff;
  font-family:'Inter',sans-serif;font-size:13px;font-weight:600;
  cursor:pointer;letter-spacing:.02em;transition:all .2s}
.btn-go:disabled{opacity:.4;cursor:not-allowed}
.btn-go:not(:disabled):hover{background:var(--navy2)}

.prog-area{display:none;margin-top:18px}
.prog-area.show{display:block}
.prog-row{display:flex;justify-content:space-between;font-size:11px;color:var(--muted);margin-bottom:7px}
.prog-track{height:4px;background:var(--border2);border-radius:4px;overflow:hidden}
.prog-bar{height:100%;background:var(--navy);width:0%;transition:width .4s ease}
.logs{margin-top:12px;max-height:100px;overflow-y:auto}
.log{padding:3px 0 3px 10px;border-left:2px solid var(--border2);
  font-size:10.5px;color:var(--muted);font-family:'JetBrains Mono',monospace;margin-bottom:2px}
.log.ok{border-color:#1E6B3C;color:#1E6B3C}
.log.err{border-color:#C00000;color:#C00000}
.log.info{border-color:var(--navy2);color:var(--navy2)}

.err-box{display:none;align-items:center;gap:10px;
  background:var(--red-bg);border:1px solid var(--red-border);
  border-radius:8px;padding:12px 16px;font-size:11.5px;color:var(--red);margin-top:14px}
.err-box.show{display:flex}

.sum-sec,.dl-sec{display:none}
.sum-sec.show,.dl-sec.show{display:block;animation:slideIn .3s ease}

.stat-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(120px,1fr));gap:12px;margin-bottom:22px}
.stat{background:var(--bg);border:1px solid var(--border);border-radius:8px;
  padding:16px 14px;border-top:3px solid var(--navy)}
.stat:nth-child(2){border-top-color:#2E6CA6}
.stat:nth-child(3){border-top-color:#1E6B3C}
.stat:nth-child(4){border-top-color:#B07D00}
.stat:nth-child(5){border-top-color:#C00000}
.stat:nth-child(6){border-top-color:#5A3B8C}
.stat-v{font-size:24px;font-weight:700;line-height:1;margin-bottom:5px;color:var(--text)}
.stat-l{font-size:9.5px;font-weight:600;text-transform:uppercase;letter-spacing:.06em;color:var(--muted)}

.meta-grid{display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:22px}
.meta-item{background:var(--bg);border:1px solid var(--border);border-radius:8px;padding:12px 14px}
.meta-k{font-size:9.5px;font-weight:600;text-transform:uppercase;letter-spacing:.06em;
  color:var(--muted);display:block;margin-bottom:4px}
.meta-v{font-size:12px;color:var(--text);word-break:break-word;line-height:1.4}
.badge{display:inline-block;padding:2px 10px;border-radius:20px;font-size:10.5px;font-weight:500}
.badge-navy{background:#EEF4FB;border:1px solid #C3D8EF;color:var(--navy2)}
.badge-green{background:var(--green-bg);border:1px solid #A8D8BC;color:var(--green)}
.badge-yellow{background:var(--yellow-bg);border:1px solid var(--yellow-border);color:var(--yellow)}
.badge-red{background:var(--red-bg);border:1px solid var(--red-border);color:var(--red)}

.ann-list{display:flex;flex-wrap:wrap;gap:8px;margin-top:4px}
.ann-tag{background:#EEF4FB;border:1px solid #C3D8EF;border-radius:6px;
  padding:4px 12px;font-size:11px;display:flex;align-items:center;gap:8px;color:var(--navy2)}
.ann-n{font-weight:700}

.tbl-wrap{overflow-x:auto;border-radius:8px;border:1px solid var(--border)}
table{width:100%;border-collapse:collapse;font-size:10.5px}
thead th{background:#EEF4FB;border-bottom:2px solid #C3D8EF;
  padding:9px 12px;text-align:left;font-size:9.5px;font-weight:600;
  letter-spacing:.05em;text-transform:uppercase;color:var(--navy);white-space:nowrap}
tbody tr{border-bottom:1px solid var(--border2)}
tbody tr:hover{background:#F8FAFC}
tbody td{padding:8px 12px;color:var(--text);max-width:180px;
  overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
td.td-dup{color:var(--yellow);font-weight:600}
td.td-sap{color:var(--red);font-weight:600}
td.td-spir{color:var(--navy);font-weight:500}
.tbl-more{text-align:center;padding:9px;color:var(--muted);
  font-size:10px;border-top:1px solid var(--border2)}

.dl-card{background:linear-gradient(135deg,#EEF4FB,#F0F7F4);
  border:1px solid #C3D8EF;border-radius:10px;padding:36px;text-align:center}
.dl-icon{font-size:48px;display:block;margin-bottom:16px;animation:bob 2.5s ease-in-out infinite}
@keyframes bob{0%,100%{transform:translateY(0)}50%{transform:translateY(-6px)}}
.dl-card h3{font-size:18px;font-weight:700;margin-bottom:8px}
.dl-card p{font-size:11.5px;color:var(--muted);margin-bottom:24px;line-height:1.6}
.btn-dl{display:inline-flex;align-items:center;gap:8px;
  padding:12px 32px;background:var(--navy);color:#fff;
  font-family:'Inter',sans-serif;font-size:13px;font-weight:600;
  border:none;border-radius:8px;cursor:pointer;text-decoration:none;
  letter-spacing:.02em;transition:background .2s}
.btn-dl:hover{background:var(--navy2)}
.btn-rst{display:inline-flex;align-items:center;gap:6px;
  padding:9px 20px;margin-top:14px;background:transparent;
  color:var(--muted);font-family:'Inter',sans-serif;font-size:11.5px;
  border:1px solid var(--border);border-radius:7px;cursor:pointer;transition:all .2s}
.btn-rst:hover{color:var(--text);border-color:var(--muted)}

.legend{display:flex;gap:16px;flex-wrap:wrap;margin-top:16px;padding-top:16px;border-top:1px solid var(--border2)}
.leg-item{display:flex;align-items:center;gap:6px;font-size:10.5px;color:var(--muted)}
.leg-swatch{width:14px;height:14px;border-radius:3px;flex-shrink:0}

::-webkit-scrollbar{width:4px;height:4px}
::-webkit-scrollbar-track{background:var(--bg)}
::-webkit-scrollbar-thumb{background:var(--border);border-radius:4px}
</style>
</head>
<body>
<header>
  <div class="hdr-inner">
    <div class="logo">SP</div>
    <div class="hdr-text">
      <h1>SPIR Extraction Tool</h1>
      <p>SPARE PARTS &amp; INTERCHANGEABILITY RECORD · AUTOMATED EXTRACTION ENGINE</p>
    </div>
    <span class="ver">v8.4 · Multi-Format</span>
  </div>
</header>

<main>
<div class="wrap">

  <div class="stepper">
    <div class="step active" id="s1"><div class="s-num" id="sn1">1</div><span class="s-lbl">Upload</span></div>
    <div class="s-line" id="sl1"></div>
    <div class="step" id="s2"><div class="s-num" id="sn2">2</div><span class="s-lbl">Process</span></div>
    <div class="s-line" id="sl2"></div>
    <div class="step" id="s3"><div class="s-num" id="sn3">3</div><span class="s-lbl">Review</span></div>
    <div class="s-line" id="sl3"></div>
    <div class="step" id="s4"><div class="s-num" id="sn4">4</div><span class="s-lbl">Download</span></div>
  </div>

  <div class="card">
    <div class="card-title">Upload SPIR File</div>
    <div class="drop-zone" id="drop">
      <span class="drop-icon">📂</span>
      <h3>Drop your SPIR Excel file here</h3>
      <p>Supports .xlsx and .xlsm · Format auto-detected on upload</p>
      <div class="fmt-pills">
        <span class="fpill">FORMAT 1 · Multi-Annexure (.xlsx)</span>
        <span class="fpill">FORMAT 2 · Single-Sheet, 1 Tag (.xlsm)</span>
        <span class="fpill">FORMAT 3 · Single-Sheet, Multi-Tag (.xlsm)</span>
        <span class="fpill">FORMAT 4 · Matrix + Continuation Sheet (.xlsx)</span>
        <span class="fpill">FORMAT 5 · Flag SPIR + Multi-Continuation (.xlsm)</span>
      </div>
      <input type="file" id="fi" accept=".xlsx,.xls,.xlsm">
    </div>
    <div class="file-sel" id="fsel">
      <span style="font-size:22px">📊</span>
      <div class="fi-info"><strong id="fname">—</strong><span id="fsize">—</span></div>
      <span class="check-icon">✓</span>
    </div>
    <div class="err-box" id="errbox"><span>⚠</span><span id="errmsg"></span></div>
    <button class="btn-go" id="pbtn" disabled onclick="go()">⚡ Extract SPIR Data</button>
    <div class="prog-area" id="prog">
      <div class="prog-row"><span id="plbl">Processing…</span><span id="ppct">0%</span></div>
      <div class="prog-track"><div class="prog-bar" id="pb2"></div></div>
      <div class="logs" id="logs"></div>
    </div>
  </div>

  <div class="sum-sec" id="sumSec">
    <div class="card">
      <div class="card-title">Extraction Summary</div>
      <div class="stat-grid" id="statGrid"></div>
      <div class="meta-grid" id="metaGrid"></div>
      <div class="card-title" id="annLabel" style="margin-top:4px">Sheet Breakdown</div>
      <div class="ann-list" id="annList"></div>
      <div class="legend">
        <div class="leg-item"><div class="leg-swatch" style="background:#D9E1F2;border:1px solid #B0BED0"></div>Equipment header row</div>
        <div class="leg-item"><div class="leg-swatch" style="background:#FFF2CC;border:1px solid #F0D060"></div>Duplicate item</div>
        <div class="leg-item"><div class="leg-swatch" style="background:#FCE4D6;border:1px solid #F0B0A0"></div>SAP Number Mismatch</div>
      </div>
    </div>
    <div class="card">
      <div class="card-title">Data Preview <span style="font-size:10px;color:var(--muted2);font-weight:400;text-transform:none;letter-spacing:0">(first 12 rows)</span></div>
      <div class="tbl-wrap">
        <table><thead id="ph"></thead><tbody id="ptb"></tbody></table>
      </div>
      <div class="tbl-more" id="tblMore"></div>
    </div>
  </div>

  <div class="dl-sec" id="dlSec">
    <div class="card">
      <div class="card-title">Download</div>
      <div class="dl-card">
        <span class="dl-icon">📥</span>
        <h3>Extraction Complete</h3>
        <p id="dlDesc">—</p>
        <a id="dlnk" class="btn-dl" href="#">⬇ Download XLSX</a>
        <br>
        <button class="btn-rst" onclick="resetTool()">↺ Process another file</button>
      </div>
    </div>
  </div>

</div>
</main>

<script>
let sf=null;
const drop=document.getElementById('drop'),fi=document.getElementById('fi');
drop.addEventListener('click',()=>fi.click());
drop.addEventListener('dragover',e=>{e.preventDefault();drop.classList.add('over')});
drop.addEventListener('dragleave',()=>drop.classList.remove('over'));
drop.addEventListener('drop',e=>{e.preventDefault();drop.classList.remove('over');if(e.dataTransfer.files[0])hf(e.dataTransfer.files[0])});
fi.addEventListener('change',e=>{if(e.target.files[0])hf(e.target.files[0])});
function hf(f){
  if(!f.name.match(/\.(xlsx|xls|xlsm)$/i)){se('Please upload .xlsx or .xlsm');return}
  ce();sf=f;
  document.getElementById('fname').textContent=f.name;
  document.getElementById('fsize').textContent=fb(f.size);
  document.getElementById('fsel').classList.add('show');
  document.getElementById('pbtn').disabled=false;
}
function fb(b){return b<1024?b+' B':b<1048576?(b/1024).toFixed(1)+' KB':(b/1048576).toFixed(2)+' MB'}
function se(m){document.getElementById('errmsg').textContent=m;document.getElementById('errbox').classList.add('show')}
function ce(){document.getElementById('errbox').classList.remove('show')}
function ss(n){
  for(let i=1;i<=4;i++){
    const s=document.getElementById('s'+i);s.classList.remove('active','done');
    if(i<n)s.classList.add('done');if(i===n)s.classList.add('active');
    if(i<4)document.getElementById('sl'+i).classList.toggle('done',i<n);
  }
}
function lg(m,c='info'){const el=document.getElementById('logs'),d=document.createElement('div');d.className='log '+c;d.textContent='› '+m;el.appendChild(d);el.scrollTop=el.scrollHeight}
function sp(p,l){document.getElementById('pb2').style.width=p+'%';document.getElementById('ppct').textContent=p+'%';document.getElementById('plbl').textContent=l}
async function go(){
  if(!sf)return;ce();ss(2);
  document.getElementById('pbtn').disabled=true;
  document.getElementById('prog').classList.add('show');
  document.getElementById('logs').innerHTML='';
  const fd=new FormData();fd.append('file',sf);
  sp(10,'Uploading…');lg('Uploading: '+sf.name);
  try{
    const res=await fetch('/extract',{method:'POST',body:fd});
    sp(80,'Building output…');
    const d=await res.json();
    if(!res.ok||d.error)throw new Error(d.error||'Server error');
    sp(100,'Done!');
    lg('Format detected: '+d.format,'info');
    lg('SPIR Type: '+(d.spir_type||'—'),'info');
    lg('Tags found: '+d.total_tags,'info');
    if(d.dup1_count>0)lg(d.dup1_count+' duplicate row(s) flagged','info');
    if(d.sap_count>0)lg(d.sap_count+' SAP mismatch row(s) flagged','info');
    lg(d.total_rows+' total rows extracted.','ok');
    rs(d);
    document.getElementById('sumSec').classList.add('show');
    document.getElementById('dlSec').classList.add('show');
    document.getElementById('dlnk').href='/download/'+d.file_id;
    document.getElementById('dlnk').download=d.filename;
    document.getElementById('dlDesc').textContent=
      d.total_rows+' rows · '+d.spare_items+' spare items · '+d.total_tags+' tags · '+(d.spir_type||'—')+' · '+d.filename;
    ss(4);
  }catch(e){lg('ERROR: '+e.message,'err');se(e.message);ss(1);document.getElementById('pbtn').disabled=false}
}
function rs(d){
  const d1=d.dup1_count,sm=d.sap_count;
  document.getElementById('statGrid').innerHTML=`
    <div class="stat"><div class="stat-v">${d.total_rows}</div><div class="stat-l">Output Rows</div></div>
    <div class="stat"><div class="stat-v">${d.spare_items}</div><div class="stat-l">Spare Items</div></div>
    <div class="stat"><div class="stat-v">${d.total_tags}</div><div class="stat-l">Eqpt Tags</div></div>
    <div class="stat"><div class="stat-v">${d1}</div><div class="stat-l">Duplicates</div></div>
    <div class="stat"><div class="stat-v">${sm}</div><div class="stat-l">SAP Mismatches</div></div>
    <div class="stat"><div class="stat-v">${d.eqpt_qty||'—'}</div><div class="stat-l">Eqpt Qty</div></div>`;
  const tv=d.spir_type?`<span class="badge badge-green">${d.spir_type}</span>`:'<span style="color:var(--muted)">—</span>';
  const fv=`<span class="badge badge-navy">${d.format}</span>`;
  const dv=d1>0?`<span class="badge badge-yellow">${d1} rows</span>`:'<span style="color:var(--green);font-size:11px;font-weight:600">✓ None</span>';
  const sv=sm>0?`<span class="badge badge-red">${sm} rows</span>`:'<span style="color:var(--green);font-size:11px;font-weight:600">✓ None</span>';
  document.getElementById('metaGrid').innerHTML=`
    <div class="meta-item"><span class="meta-k">SPIR Number</span><div class="meta-v" style="color:var(--navy);font-weight:600">${d.spir_no||'—'}</div></div>
    <div class="meta-item"><span class="meta-k">Format</span><div class="meta-v">${fv}</div></div>
    <div class="meta-item"><span class="meta-k">SPIR Type</span><div class="meta-v">${tv}</div></div>
    <div class="meta-item"><span class="meta-k">Duplicates (Type 1)</span><div class="meta-v">${dv}</div></div>
    <div class="meta-item"><span class="meta-k">SAP Mismatch (Type 2)</span><div class="meta-v">${sv}</div></div>
    <div class="meta-item"><span class="meta-k">Manufacturer</span><div class="meta-v">${d.manufacturer||'—'}</div></div>`;
  const al=document.getElementById('annLabel'),aa=document.getElementById('annList');
  if(!Object.keys(d.annexure_stats).length){al.style.display='none';aa.style.display='none'}
  else{al.style.display='flex';aa.style.display='flex';
    aa.innerHTML=Object.entries(d.annexure_stats).map(([k,v])=>`<div class="ann-tag"><span>${k}</span><span class="ann-n">${v} tags</span></div>`).join('')}
  const PCOLS=['SPIR NO','TAG NO','EQPT MODEL','EQPT QTY','ITEM NUMBER','DESCRIPTION OF PARTS','NEW DESCRIPTION OF PARTS','MANUFACTURER PART NUMBER','SAP NUMBER','UNIT PRICE','DUPLICATE ID','SHEET'];
  document.getElementById('ph').innerHTML='<tr>'+PCOLS.map(c=>`<th>${c}</th>`).join('')+'</tr>';
  const ac=d.preview_cols,ix=PCOLS.map(c=>ac.indexOf(c));
  const dupi=ac.indexOf('DUPLICATE ID'),spiri=ac.indexOf('SPIR NO');
  document.getElementById('ptb').innerHTML=d.preview_rows.map(r=>'<tr>'+ix.map(i=>{
    const v=r[i]??'';
    if(i===dupi){
      if(String(v).startsWith('Duplicate'))return`<td class="td-dup">${v}</td>`;
      if(String(v)==='SAP NUMBER MISMATCH')return`<td class="td-sap">${v}</td>`;
    }
    if(i===spiri&&v)return`<td class="td-spir">${v}</td>`;
    return`<td title="${v}">${v}</td>`;
  }).join('')+'</tr>').join('');
  const more=d.total_rows-12;
  document.getElementById('tblMore').textContent=more>0?`+ ${more} more rows in the downloaded file`:'';
}
function resetTool(){
  sf=null;fi.value='';
  ['fsel','prog','sumSec','dlSec'].forEach(id=>document.getElementById(id).classList.remove('show'));
  document.getElementById('pbtn').disabled=true;
  document.getElementById('logs').innerHTML='';
  document.getElementById('pb2').style.width='0%';
  ce();ss(1);
}
</script>
</body>
</html>"""


# ─── ENTRY POINT ──────────────────────────────────────────────────────────────

if __name__ == '__main__':
    port = 5050
    print(f'\n  ✦ SPIR Extraction Tool v8.4  →  http://localhost:{port}\n')
    print('  Formats supported:')
    print('    FORMAT 1 — Multi-Annexure (.xlsx)')
    print('    FORMAT 2 — Single-Sheet, 1 Tag (.xlsm)')
    print('    FORMAT 3 — Single-Sheet, Multi-Tag (.xlsm)')
    print('    FORMAT 4 — Matrix SPIR + Single Continuation Sheet (.xlsx)')
    print('    FORMAT 5 — Flag SPIR + Multiple Continuation Sheets (.xlsm)\n')
    print('  New Description = Description + Part Number + Supplier (non-empty parts joined)\n')
    threading.Timer(1.2, lambda: webbrowser.open(f'http://localhost:{port}')).start()
    app.run(port=port, debug=False)
